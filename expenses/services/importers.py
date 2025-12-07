import csv
import io
import re
from datetime import datetime, date
from decimal import Decimal

import openpyxl

from ..models import Expense, ExpenseCategory


def _normalize_amount(raw) -> Decimal:
    """
    Приводим сумму к Decimal, поддерживаем как строки, так и числа.
    """
    if raw is None:
        return Decimal("0")

    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw))

    text = (
        str(raw)
        .replace("\xa0", "")
        .replace(" ", "")
        .replace("+", "")
        .replace("−", "-")
        .replace(",", ".")
    )
    if not text:
        return Decimal("0")
    return Decimal(text)

MONTHS_RU = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "мая": 5,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "сент": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}


def _parse_sber_date(raw) -> date | None:
    """
    Поддерживаем форматы типа:
    - '01 дек. 2025, ...'
    - '01 дек. 2025'
    - '01.12.2025'
    - уже готовый datetime/date.
    """
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw

    if raw is None:
        return None

    s = str(raw).strip()

    # Сначала пробуем цифровой формат с точками
    main_part = s.split(",")[0].strip()
    try:
        return datetime.strptime(main_part, "%d.%m.%Y").date()
    except Exception:
        pass

    # Формат с русским месяцем: '01 дек. 2025'
    m = re.search(r"(\d{1,2})\s+([А-Яа-яЁё]{3,})\.?,?\s+(\d{4})", s)
    if m:
        day = int(m.group(1))
        mon_text = m.group(2).lower()
        year = int(m.group(3))

        month_num = None
        for key, value in MONTHS_RU.items():
            if mon_text.startswith(key):
                month_num = value
                break

        if month_num:
            try:
                return date(year, month_num, day)
            except ValueError:
                return None

    return None



# ---------- T-Bank CSV ----------

def import_tbank_csv(file_obj, user):
    """
    Import expenses from T-Bank (Tinkoff) CSV export.
    Only rows with negative 'Сумма операции' are treated as expenses.
    Returns (created_count, total_amount).
    """
    content = file_obj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content), delimiter=";")

    created = 0
    total = Decimal("0")

    for row in reader:
        status = row.get("Статус")
        if status != "OK":
            continue

        amount_raw = row.get("Сумма операции") or row.get("Сумма платежа") or ""
        if not amount_raw:
            continue

        amount = _normalize_amount(amount_raw)
        if amount >= 0:
            # пополнения / возвраты не считаем расходами
            continue
        amount = -amount  # храним как положительную сумму расхода

        currency = (row.get("Валюта операции") or "RUB").strip()
        date_str = (row.get("Дата операции") or "").split()[0]
        try:
            date = datetime.strptime(date_str, "%d.%m.%Y").date()
        except ValueError:
            continue

        category_name = (row.get("Категория") or "Uncategorized").strip()
        description = (row.get("Описание") or "").strip()

        category, _ = ExpenseCategory.objects.get_or_create(
            user=user,
            name=category_name,
        )

        Expense.objects.create(
            user=user,
            category=category,
            date=date,
            description=description[:255] or category_name,
            amount=amount,
            bank="T-Bank",
            currency=currency,
        )
        created += 1
        total += amount

    return created, total


# ---------- Sber XLSX ----------

def import_sber_xlsx(file_obj, user):
    """
    Import expenses from Sberbank XLSX statement.

    Ожидаем шапку вида:
    Номер | Дата | Тип операции | Категория | Сумма | Валюта | Сумма в руб | Описание | ...

    Но названия столбцов ищем по подстрокам, чтобы быть устойчивыми к мелким отличиям.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # --- 1. Берём первую строку как заголовок (как в твоём примере) --- #
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [(str(c).strip() if c is not None else "") for c in first_row]

    def find_idx(keywords):
        """
        Находит индекс колонки, в названии которой есть одно из слов из keywords.
        keywords: список подстрок в нижнем регистре.
        """
        lower_headers = [h.lower() for h in headers]
        for i, h in enumerate(lower_headers):
            for kw in keywords:
                if kw in h:
                    return i
        return -1

    idx_date = find_idx(["дата"])
    idx_category = find_idx(["категор"])
    idx_amount = find_idx(["сумма"])          # попадает в E: 'Сумма'
    idx_amount_rub = find_idx(["сумма в руб"])  # попадает в G: 'Сумма в руб'
    idx_currency = find_idx(["валют"])
    idx_description = find_idx(["описан"])
    # idx_type = find_idx(["тип операц"])  # можно использовать позже

    if idx_date < 0 or idx_amount < 0:
        # формат неожиданный — просто выходим
        return 0, Decimal("0")

    created = 0
    total = Decimal("0")

    # --- 2. Идём по строкам данных --- #
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        # Сумма
        raw_amount = row[idx_amount]
        amount = _normalize_amount(raw_amount)

        if amount == 0:
            continue

        # В выписке Сбера расходы обычно положительные.
        # Если вдруг отрицательные — берём модуль.
        if amount < 0:
            amount = -amount

        # Дата
        raw_date = row[idx_date]
        date_value = _parse_sber_date(raw_date)
        if not date_value:
            continue

        # Категория
        raw_cat = row[idx_category] if idx_category >= 0 else ""
        category_name = (str(raw_cat) if raw_cat is not None else "").strip() or "Uncategorized"
        category_name_lower = category_name.lower()

        # Отсекаем очевидные переводы/пополнения
        if any(bad in category_name_lower for bad in ["перевод", "пополн", "зачислен"]):
            continue

        # Описание
        raw_desc = row[idx_description] if idx_description >= 0 else ""
        description = (str(raw_desc) if raw_desc is not None else "").strip() or category_name

        # Валюта
        raw_cur = row[idx_currency] if idx_currency >= 0 else ""
        currency = (str(raw_cur) if raw_cur is not None else "").strip() or "RUB"

        # Если есть "Сумма в руб" и валюта не RUB — используем её
        if currency != "RUB" and idx_amount_rub >= 0 and row[idx_amount_rub] is not None:
            amount_rub = _normalize_amount(row[idx_amount_rub])
            if amount_rub != 0:
                amount = abs(amount_rub)
                currency = "RUB"

        category, _ = ExpenseCategory.objects.get_or_create(
            user=user,
            name=category_name,
        )

        Expense.objects.create(
            user=user,
            category=category,
            date=date_value,
            description=description[:255],
            amount=amount,
            bank="Sberbank",
            currency=currency,
        )

        created += 1
        total += amount

    return created, total
